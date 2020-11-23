import re

import openpyxl as xl
import datetime

from algorthim.findSector import findSectorAlgor
from database import transaction_mgt as tm
from database import credit_mgt as cm

from database.redis.myConnRedis import business_names_db

class Table:
    def __init__(self, start_idx, credit_card):
        self.start_idx = start_idx
        self.credit_card = credit_card

    def __str__(self) -> str:
        return "start_idx: " + str(self.start_idx) + " credit_card: " + str(self.credit_card)


def readExcelFile(filename: str, user_id):
    wb = xl.load_workbook(filename)
    ws = wb['Transactions']

    tables = []
    cm.create_credit_table()
    for row in ws['B']:
        if row.value == "מועד חיוב":
            start_idx = row.row
            credit_card = ws.cell(start_idx, 1).value
            credit_card = getCreditNum(credit_card)
            cm.add_credit(user_id, credit_card)
            table = Table(start_idx, credit_card)
            tables.append(table)

    for i in range(len(tables)):
        if i == len(tables) - 1:  # last table
            last_row = ws.max_row
            extractTable(ws, tables[i].start_idx, last_row, tables[i].credit_card)
        else:
            extractTable(ws, tables[i].start_idx, tables[i + 1].start_idx, tables[i].credit_card)

    business_names_db.save()


def extractTable(ws, start_idx, end_idx, credit_card):
    for row in ws.iter_rows(min_row=start_idx, max_row=end_idx):
        date = row[0].value
        try:
            date = datetime.datetime.strptime(date, '%d/%m/%Y')  # convert to datetime format
            business_name = row[1].value
            sector = findSectorAlgor(business_name)
            amount = row[2].value
            tm.add_transaction(business_name, date, sector, amount, credit_card)
        except:
            pass


def getCreditNum(str):
    x = re.search("\d\d\d\d", str)
    return x.group()

